import datetime
from django.http import JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse_lazy
from django.views import View
from django.views.generic import ListView, CreateView, DetailView, UpdateView
# from estoque.forms import EntradaForm
from estoque.models import EntradaEstoque, Estoque, EstoqueImei, ProdutoEntrada
from django.contrib import messages
from django.db.models import Q
from django.contrib.auth.mixins import PermissionRequiredMixin
from django.contrib.auth.decorators import login_required, permission_required
from produtos.models import Produto
from vendas.models import Loja
from .forms import EntradaEstoqueForm, EstoqueImeiForm, ProdutoEntradaForm, ProdutoEntradaFormSet, ProdutoEntradaEditFormSet, EstoqueImeiEditForm
from vendas.views import BaseView
from vendas.models import Venda
from produtos.models import TipoProduto
from .models import Fornecedor
from django.db.models import Sum


class EstoqueListView(BaseView, PermissionRequiredMixin, ListView):
    model = Estoque
    template_name = 'estoque/estoque_list.html'
    context_object_name = 'produtos'
    permission_required = 'estoque.view_estoque'
    paginate_by = 10
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        loja_id = self.request.session.get('loja_id')
        loja = get_object_or_404(Loja, pk=loja_id)
        tipos_produtos = TipoProduto.objects.all()
        context['loja_id'] = loja_id
        context['tipos'] = tipos_produtos
        return context
    
    def get_queryset(self):
        loja_id = self.request.session.get('loja_id')
        loja = get_object_or_404(Loja, pk=loja_id)
        query = super().get_queryset().filter(loja=loja)
        search = self.request.GET.get('search', None)
        if search:
            query = query.filter(produto__nome__icontains=search)
            
        return query

class EstoqueUpdateView(PermissionRequiredMixin, UpdateView):
    model = Estoque
    fields = ['quantidade_disponivel']
    template_name = 'estoque/estoque_edit.html'
    success_url = reverse_lazy('estoque:estoque_list')
    permission_required = 'estoque.change_estoque'

    def form_valid(self, form):
        estoque = form.save(commit=False)
        estoque.save(user=self.request.user)
        messages.success(self.request, 'Estoque atualizado com sucesso!')
        return redirect(self.success_url)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        produto = self.object.produto
        context['produto'] = produto
        return context

class EntradaListView(BaseView, PermissionRequiredMixin, ListView):
    model = EntradaEstoque
    template_name = 'estoque/estoque_entrada_list.html'
    context_object_name = 'entradas'
    permission_required = 'estoque.view_entradaestoque'
    paginate_by = 10
    
    def get_queryset(self):
        query = EntradaEstoque.objects.all()
        loja_id = self.request.session.get('loja_id')
        loja = get_object_or_404(Loja, pk=loja_id)
        search = self.request.GET.get('search', None)
        liberada = self.request.GET.get('liberada', None)
        
        if liberada:
            if liberada == 'true':
                query = query.filter(venda_liberada=True)
            else:
                query = query.filter(venda_liberada=False)
        
        if not self.request.user.has_perm('estoque.can_view_all_imei'):
            query = query.filter(loja=loja)
            
        if search:
            query = query.filter(Q(numero_nota__icontains=search))
            
        return query.order_by('-criado_em')
    
    
class EntradaDetailView(PermissionRequiredMixin, DetailView):
    model = EntradaEstoque
    template_name = 'estoque/estoque_entrada_detail.html'
    context_object_name = 'entrada'
    permission_required = 'estoque.view_entradaestoque'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context
    
    
class EntradaUpdateView(PermissionRequiredMixin, UpdateView):
    # levar em consideração que pode ter uma venda e diminuir o estoque
    model = EntradaEstoque
    form_class = EntradaEstoqueForm
    template_name = 'estoque/estoque_form_edit.html'
    success_url = reverse_lazy('estoque:estoque_list')
    permission_required = 'estoque.change_entradaestoque'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        loja_id = self.request.session.get('loja_id')
        if self.request.POST:
            context['formset'] = ProdutoEntradaEditFormSet(self.request.POST, form_kwargs={'loja': loja_id})
        else:
            context['formset'] = ProdutoEntradaEditFormSet(queryset=ProdutoEntrada.objects.filter(entrada=self.object), form_kwargs={'loja': loja_id})
        return context
    
    def form_valid(self, form):
        context = self.get_context_data()
        formset = context['formset']
        loja_id = self.request.session.get('loja_id')
        loja = get_object_or_404(Loja, pk=loja_id)
        if form.is_valid() and formset.is_valid():
            entrada_estoque = form.save(commit=False)
            produtos = formset.save(commit=False)
            # verificar se excluiu algum produto
            for produto in formset.deleted_objects:
                produto.delete()

            for produto in produtos:
                produto.entrada = entrada_estoque
                produto.loja = loja
                produto.save(user=self.request.user)
                
                # Se o produto for serializado, salve os IMEIs na tabela EstoqueImei
                if produto.imei:  # Presumindo que o IMEI é obrigatório
                    estoque_imei = EstoqueImei.objects.create(
                        produto=produto.produto,
                        imei=produto.imei,
                        produto_entrada=produto,
                        loja=loja
                    )
                    estoque_imei.save(user=self.request.user)

            #verificar se a entrada não esta vazia
            if not entrada_estoque.produtos.all():
                entrada_estoque.delete()
                messages.error(self.request, 'Entrada Excluída, pois não possui produtos.')
                return redirect(self.success_url)

            entrada_estoque.save(user=self.request.user)
            messages.success(self.request, 'Entrada de estoque atualizada com sucesso!')
            return redirect(self.success_url)
        else:
            return self.form_invalid(form)


class AdicionarEntradaEstoqueView(PermissionRequiredMixin, CreateView):
    model = EntradaEstoque
    form_class = EntradaEstoqueForm
    template_name = 'estoque/estoque_form.html'
    success_url = reverse_lazy('estoque:estoque_list')
    permission_required = 'estoque.add_entradaestoque'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        loja_id = self.request.session.get('loja_id')
        if self.request.POST:
            context['formset'] = ProdutoEntradaFormSet(self.request.POST, form_kwargs={'loja': loja_id})
        else:
            context['formset'] = ProdutoEntradaFormSet(queryset=ProdutoEntrada.objects.none(), form_kwargs={'loja': loja_id})
        return context

    def form_valid(self, form):
        context = self.get_context_data()
        formset = context['formset']
        loja_id = self.request.session.get('loja_id')
        loja = get_object_or_404(Loja, pk=loja_id)
        if form.is_valid() and formset.is_valid():
            entrada_estoque = form.save(commit=False)
            entrada_estoque.loja = loja
            entrada_estoque.save(user=self.request.user)
            produtos = formset.save(commit=False)

            for produto in produtos:
                produto.entrada = entrada_estoque
                produto.loja = loja
                produto.save(user=self.request.user)
                
                if produto.imei: 
                    estoque_imei = EstoqueImei(
                        produto=produto.produto,
                        imei=produto.imei,
                        produto_entrada=produto,
                        loja=loja
                    )
                    estoque_imei.save(user=self.request.user)
            
            messages.success(self.request, 'Entrada de estoque criada com sucesso!')
            return redirect(self.success_url)
        else:
            return self.form_invalid(form)
        


@permission_required('estoque.can_liberar_venda', raise_exception=True)
def liberar_entrada(request, pk):
    entrada = get_object_or_404(EntradaEstoque, pk=pk)

    if entrada.venda_liberada:
        messages.info(request, "Esta Nota de entrada já está liberada para venda.")
    else:
        entrada.venda_liberada = True
        entrada.save(update_fields=['venda_liberada'])
        messages.success(request, "Nota de entrada liberada para venda com sucesso.")

    return redirect('estoque:entrada_detail', pk=pk)


class EstoqueImeiListView(BaseView, PermissionRequiredMixin, ListView):
    model = EstoqueImei
    template_name = 'estoque/estoque_imei_list.html'
    context_object_name = 'produtos'
    permission_required = 'estoque.view_estoqueimei'
    paginate_by = 10
    ordering = ['-id']
    
    def get_queryset(self):
        loja_id = self.request.session.get('loja_id')
        loja = get_object_or_404(Loja, pk=loja_id)
        query = super().get_queryset().filter(loja=loja)
        
        search = self.request.GET.get('search', None)
        if search:
            query = query.filter(Q(imei__icontains=search)|Q(produto__nome__icontains=search))
            
        return query
    

class EstoqueImeiUpdateView(PermissionRequiredMixin, UpdateView):
    model = EstoqueImei
    form_class = EstoqueImeiEditForm
    template_name = 'estoque/estoque_imei_form_edit.html'
    success_url = reverse_lazy('estoque:estoque_imei_list')
    permission_required = 'estoque.change_estoqueimei'

    def form_valid(self, form):
        estoque = form.save(commit=False)
        estoque.save(user=self.request.user)
        messages.success(self.request, 'Estoque atualizado com sucesso!')
        return redirect(self.success_url)

class FornecedorListView(PermissionRequiredMixin, ListView):
    model = Fornecedor
    template_name = 'fornecedor/fornecedor_list.html'
    permission_required = 'estoque.view_fornecedor'
    context_object_name = 'items'
    paginate_by = 10

    def get_queryset(self):
        query = super().get_queryset()
        search = self.request.GET.get('search', None)
        if search:
            query = query.filter(nome__icontains=search)
        return query

def cancelar_imei(request, id):
    imei = get_object_or_404(EstoqueImei, pk=id)

    if not request.user.has_perm('estoque.change_estoqueimei'):
        messages.error(request, 'Você não tem permissão para cancelar este IMEI.')
        return redirect('estoque:estoque_imei_list')
    
    if imei.cancelado:
        messages.error(request, 'Este IMEI já está cancelado.')
    else:
        imei.cancelado = True
        imei.save(user=request.user)
        messages.success(request, 'IMEI cancelado com sucesso.')
    return redirect('estoque:estoque_imei_list')

@login_required
def check_produtos(request, produto_id):
    produto = get_object_or_404(Produto, pk=produto_id)
    if produto.tipo.numero_serial:
        return JsonResponse({'serializado': True})
    else:
        return JsonResponse({'serializado': False})
    
    
class EstoqueImeiSearchView(View):
    def get(self, request, *args, **kwargs):
        term = request.GET.get('term', '')
        user = request.user
        produto_id = request.GET.get('produto_id', None)
        loja_id = self.request.session.get('loja_id')
        loja = get_object_or_404(Loja, pk=loja_id)
        queryset = EstoqueImei.objects.filter(vendido=False, produto_entrada__entrada__venda_liberada=True, cancelado=False).filter(
            Q(imei__icontains=term) | Q(produto__nome__icontains=term) | Q(loja__nome__icontains=term)
        )
        
        if not user.has_perm('estoque.can_view_all_imei'):
            queryset = queryset.filter(loja=loja)
        
        if produto_id:
            queryset = queryset.filter(produto_id=produto_id)
        
        results = []
        for imei in queryset:
            if not user.has_perm('estoque.can_view_all_imei'):
                text = f'{imei.imei} - {imei.produto.nome}'
            else:
                text = f'{imei.imei} - {imei.produto.nome} | {imei.produto.loja.nome}'
            results.append({
                'id': imei.id,
                'text': text
            })
        return JsonResponse({'results': results})
    
class EstoqueImeiSearchEditView(View):
    def get(self, request, *args, **kwargs):
        term = request.GET.get('term', '')
        produto_id = request.GET.get('produto_id', None)
        loja_id = self.request.session.get('loja_id')
        venda_id = self.request.session.get('venda_id')
        loja = get_object_or_404(Loja, pk=loja_id)
        queryset = EstoqueImei.objects.filter(
            Q(imei__icontains=term) | Q(produto__nome__icontains=term)
        ).filter(loja=loja).filter(vendido=False, cancelado=False)
        if produto_id:
            queryset = queryset.filter(produto_id=produto_id)
        results = []
        for imei in queryset:
            results.append({
                'id': imei.imei,
                'text': f'{imei.imei} - {imei.produto.nome}'
            })

        if venda_id:
            venda = get_object_or_404(Venda, pk=venda_id)
            for item in venda.itens_venda.all():
                if item.imei:
                    results.append({
                        'id': item.imei,
                        'text': f'{item.imei} - {item.produto.nome}'
                    })
        return JsonResponse({'results': results})
    
def inventario_estoque_pdf (request):
    loja = get_object_or_404(Loja, pk=request.session.get('loja_id'))
    tipo = request.GET.get('tipo', None)
    produtos = Estoque.objects.filter(loja=loja).filter(quantidade_disponivel__gt=0)

    if tipo:
        produtos = produtos.filter(produto__tipo_id=tipo)
    
    quantidade_total = produtos.aggregate(total=Sum('quantidade_disponivel'))['total']
    custo_medio_total = 0
    preco_medio_total = 0

    for produto in produtos:
        preco_medio_total += float(produto.preco_medio()) * produto.quantidade_disponivel
        custo_medio_total  += float(produto.preco_medio_custo()) * produto.quantidade_disponivel

    context = {
        'produtos': produtos,
        'loja': loja,
        'quantidade_total': quantidade_total,
        'custo_medio_total': f'{custo_medio_total:.2f}',
        'preco_medio_total': f'{preco_medio_total:.2f}',
    }

    return render(request, "estoque/folha_estoque.html", context)

def inventario_estoque_imei_pdf (request):
    loja = get_object_or_404(Loja, pk=request.session.get('loja_id'))
    produtos = EstoqueImei.objects.filter(loja=loja).filter(vendido=False)
    
    context = {
        'produtos': produtos,
        'loja': loja,
        'data_hoje': datetime.datetime.now()
    }

    return render(request, "estoque/folha_estoque_imei.html", context)

def inventario_estoque_imei_excel(request):
    """Gera relatório Excel do estoque de IMEI com métricas financeiras e estratégicas"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse
    from decimal import Decimal
    
    loja = get_object_or_404(Loja, pk=request.session.get('loja_id'))
    
    # Buscar todos os IMEIs da loja
    imeis = EstoqueImei.objects.filter(loja=loja).select_related(
        'produto', 'produto_entrada', 'produto_entrada__entrada'
    ).prefetch_related('produto__tipo')
    
    # Criar workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Estoque IMEI"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    title_font = Font(bold=True, size=14)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # Título
    ws.merge_cells('A1:N1')
    ws['A1'] = f"RELATÓRIO DE ESTOQUE IMEI - {loja.nome}"
    ws['A1'].font = title_font
    ws['A1'].alignment = center_alignment
    
    # Data do relatório
    ws.merge_cells('A2:N2')
    ws['A2'] = f"Data de geração: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws['A2'].alignment = center_alignment
    
    # Cabeçalhos da tabela
    headers = [
        'ID', 'Produto', 'IMEI', 'Tipo Produto', 'Vendido', 'App Instalado', 
        'Cancelado', 'Data Venda', 'ID Venda', 'Número Nota', 'Custo Unitário',
        'Preço Venda', 'Margem', 'Status'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_alignment
    
    # Dados dos IMEIs
    row = 5
    total_custo = Decimal('0.00')
    total_venda = Decimal('0.00')
    total_margem = Decimal('0.00')
    imeis_vendidos = 0
    imeis_disponiveis = 0
    imeis_cancelados = 0
    
    for imei in imeis:
        # Calcular custo e preço
        custo_unitario = Decimal('0.00')
        preco_venda = Decimal('0.00')
        
        if imei.produto_entrada and imei.produto_entrada.custo_unitario:
            custo_unitario = imei.produto_entrada.custo_unitario
        
        if imei.produto.valor_repasse_logista:
            preco_venda = imei.produto.valor_repasse_logista
        
        margem = preco_venda - custo_unitario if preco_venda > 0 and custo_unitario > 0 else Decimal('0.00')
        
        # Status do IMEI
        if imei.cancelado:
            status = "Cancelado"
            imeis_cancelados += 1
        elif imei.vendido:
            status = "Vendido"
            imeis_vendidos += 1
        else:
            status = "Disponível"
            imeis_disponiveis += 1
            total_custo += custo_unitario
            total_venda += preco_venda
            total_margem += margem
        
        # Preencher linha
        ws.cell(row=row, column=1, value=imei.id).border = border
        ws.cell(row=row, column=2, value=imei.produto.nome).border = border
        ws.cell(row=row, column=3, value=imei.imei).border = border
        ws.cell(row=row, column=4, value=imei.produto.tipo.nome if imei.produto.tipo else "-").border = border
        ws.cell(row=row, column=5, value="Sim" if imei.vendido else "Não").border = border
        ws.cell(row=row, column=6, value="Sim" if imei.aplicativo_instalado else "Não").border = border
        ws.cell(row=row, column=7, value="Sim" if imei.cancelado else "Não").border = border
        ws.cell(row=row, column=8, value=imei.data_venda.strftime('%d/%m/%Y') if imei.data_venda else "-").border = border
        ws.cell(row=row, column=9, value=imei.id_venda or "-").border = border
        ws.cell(row=row, column=10, value=imei.numero_nota or "-").border = border
        ws.cell(row=row, column=11, value=float(custo_unitario)).border = border
        ws.cell(row=row, column=12, value=float(preco_venda)).border = border
        ws.cell(row=row, column=13, value=float(margem)).border = border
        ws.cell(row=row, column=14, value=status).border = border
        
        row += 1
    
    # Métricas Financeiras e Estratégicas
    row_metrics = row + 2
    
    # Título das métricas
    ws.merge_cells(f'A{row_metrics}:N{row_metrics}')
    ws[f'A{row_metrics}'] = "MÉTRICAS FINANCEIRAS E ESTRATÉGICAS"
    ws[f'A{row_metrics}'].font = title_font
    ws[f'A{row_metrics}'].alignment = center_alignment
    
    row_metrics += 2
    
    # Métricas de Quantidade
    ws.cell(row=row_metrics, column=1, value="QUANTITATIVAS").font = Font(bold=True)
    ws.cell(row=row_metrics, column=2, value="Valor").font = Font(bold=True)
    ws.cell(row=row_metrics, column=3, value="Percentual").font = Font(bold=True)
    
    total_imeis = imeis.count()
    row_metrics += 1
    
    ws.cell(row=row_metrics, column=1, value="Total de IMEIs")
    ws.cell(row=row_metrics, column=2, value=total_imeis)
    ws.cell(row=row_metrics, column=3, value="100%")
    
    row_metrics += 1
    ws.cell(row=row_metrics, column=1, value="IMEIs Disponíveis")
    ws.cell(row=row_metrics, column=2, value=imeis_disponiveis)
    ws.cell(row=row_metrics, column=3, value=f"{(imeis_disponiveis/total_imeis*100):.1f}%" if total_imeis > 0 else "0%")
    
    row_metrics += 1
    ws.cell(row=row_metrics, column=1, value="IMEIs Vendidos")
    ws.cell(row=row_metrics, column=2, value=imeis_vendidos)
    ws.cell(row=row_metrics, column=3, value=f"{(imeis_vendidos/total_imeis*100):.1f}%" if total_imeis > 0 else "0%")
    
    row_metrics += 1
    ws.cell(row=row_metrics, column=1, value="IMEIs Cancelados")
    ws.cell(row=row_metrics, column=2, value=imeis_cancelados)
    ws.cell(row=row_metrics, column=3, value=f"{(imeis_cancelados/total_imeis*100):.1f}%" if total_imeis > 0 else "0%")
    
    row_metrics += 2
    
    # Métricas Financeiras
    ws.cell(row=row_metrics, column=1, value="FINANCEIRAS").font = Font(bold=True)
    ws.cell(row=row_metrics, column=2, value="Valor (R$)").font = Font(bold=True)
    
    row_metrics += 1
    ws.cell(row=row_metrics, column=1, value="Valor Total em Estoque (Custo)")
    ws.cell(row=row_metrics, column=2, value=float(total_custo))
    
    row_metrics += 1
    ws.cell(row=row_metrics, column=1, value="Valor Total em Estoque (Venda)")
    ws.cell(row=row_metrics, column=2, value=float(total_venda))
    
    row_metrics += 1
    ws.cell(row=row_metrics, column=1, value="Margem Total Potencial")
    ws.cell(row=row_metrics, column=2, value=float(total_margem))
    
    row_metrics += 1
    margem_percentual = (total_margem / total_venda * 100) if total_venda > 0 else 0
    ws.cell(row=row_metrics, column=1, value="Margem Percentual Média")
    ws.cell(row=row_metrics, column=2, value=f"{margem_percentual:.1f}%")
    
    row_metrics += 2
    
    # Métricas Estratégicas
    ws.cell(row=row_metrics, column=1, value="ESTRATÉGICAS").font = Font(bold=True)
    ws.cell(row=row_metrics, column=2, value="Valor").font = Font(bold=True)
    
    row_metrics += 1
    # Taxa de conversão (vendidos / total)
    taxa_conversao = (imeis_vendidos / total_imeis * 100) if total_imeis > 0 else 0
    ws.cell(row=row_metrics, column=1, value="Taxa de Conversão (Vendidos/Total)")
    ws.cell(row=row_metrics, column=2, value=f"{taxa_conversao:.1f}%")
    
    row_metrics += 1
    # Taxa de cancelamento
    taxa_cancelamento = (imeis_cancelados / total_imeis * 100) if total_imeis > 0 else 0
    ws.cell(row=row_metrics, column=1, value="Taxa de Cancelamento")
    ws.cell(row=row_metrics, column=2, value=f"{taxa_cancelamento:.1f}%")
    
    row_metrics += 1
    # Valor médio por IMEI
    valor_medio_imei = total_venda / imeis_disponiveis if imeis_disponiveis > 0 else 0
    ws.cell(row=row_metrics, column=1, value="Valor Médio por IMEI Disponível")
    ws.cell(row=row_metrics, column=2, value=f"R$ {valor_medio_imei:.2f}")
    
    row_metrics += 1
    # Margem média por IMEI
    margem_media_imei = total_margem / imeis_disponiveis if imeis_disponiveis > 0 else 0
    ws.cell(row=row_metrics, column=1, value="Margem Média por IMEI Disponível")
    ws.cell(row=row_metrics, column=2, value=f"R$ {margem_media_imei:.2f}")
    
    # Ajustar largura das colunas
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # Configurar resposta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="relatorio_estoque_imei_{loja.nome}_{datetime.datetime.now().strftime("%Y%m%d_%H%M")}.xlsx"'
    
    wb.save(response)
    return response

class FolhaNotaEntradaView(View):
    def get(self, request, *args, **kwargs):
        loja = get_object_or_404(Loja, pk=request.session.get('loja_id'))
        entrada = get_object_or_404(EntradaEstoque, pk=kwargs['pk'])
        produtos = ProdutoEntrada.objects.filter(entrada=entrada)
        
        context = {
            'produtos': produtos,
            'loja': loja,
            'entrada': entrada,
            'data_hoje': datetime.datetime.now()
        }

        return render(request, "estoque/folha_entrada.html", context)
    
def buscar_imei_por_produto(request):
    if request.method == 'GET':
        produto_id = request.GET.get('produto_id')
        imei_antigo = request.GET.get('imei')

        if produto_id:
            imeis = EstoqueImei.objects.filter(produto_id=produto_id, vendido=False, cancelado=False).exclude(imei=imei_antigo) if imei_antigo else EstoqueImei.objects.filter(produto_id=produto_id, vendido=False)
            imeis_list = [{'id': imei.id, 'imei': f'{imei.imei} - {imei.produto.nome}'} for imei in imeis]
            return JsonResponse({'imeis': imeis_list})
        
    return JsonResponse({'error': 'Produto não encontrado'}, status=404)